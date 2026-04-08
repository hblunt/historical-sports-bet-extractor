"""
Sports Bet Extractor — Desktop GUI for fetching Australian horse racing results.
Fetches winners and odds from the public TAB API and exports to .xlsx.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
import threading
import os
import subprocess
import platform

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ── API / Data Logic ────────────────────────────────────────────────────────

BASE_URL = "https://api.beta.tab.com.au/v1"
HEADERS = {
    "Accept": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Connection": "close",
}

JURISDICTIONS = ["QLD", "NSW", "VIC", "SA", "WA", "TAS", "ACT", "NT"]

RACE_TYPES = {
    "Thoroughbreds": "R",
    "Harness": "H",
    "Greyhounds": "G",
    "All": None,
}

MAIN_VENUES = [
    "Randwick", "Flemington", "Caulfield", "Eagle Farm", "Doomben", "Gold Coast",
]

VENUE_FILTERS = {
    "All venues": None,
    "Main races only": MAIN_VENUES,
}

ALL_COLUMNS = [
    "Date", "Venue", "Race", "Race Name", "Distance",
    "Winner", "No.", "Barrier", "Jockey", "Trainer",
    "Fixed Win", "Tote Win",
]

# Columns included by default
DEFAULT_COLUMNS = set(ALL_COLUMNS)


def fetch_meetings(date_str, jurisdiction, race_type_filter=None):
    """Fetch all race meetings for a given date and jurisdiction."""
    url = f"{BASE_URL}/historical-results-service/{jurisdiction}/racing/{date_str}"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    meetings = data.get("meetings", [])
    if race_type_filter:
        meetings = [m for m in meetings if m.get("raceType") == race_type_filter]
    return meetings


def fetch_race_detail(race_link):
    """Fetch full detail for a single race via its _links.self URL."""
    resp = requests.get(race_link, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return resp.json()


def extract_runner_data(runner, race_data):
    """Extract data for a single runner."""
    fixed_odds = runner.get("fixedOdds", {})
    win_price = fixed_odds.get("returnWin")

    pari = runner.get("parimutuel", {})
    pari_win = pari.get("returnWin")

    return {
        "runner_name": runner.get("runnerName", "Unknown"),
        "runner_number": runner.get("runnerNumber"),
        "barrier": runner.get("barrierNumber"),
        "fixed_win": win_price,
        "pari_win": pari_win,
        "jockey": runner.get("riderDriverName", ""),
        "trainer": runner.get("trainerName", ""),
    }


def extract_results(race_data, winners_only=True):
    """Extract result info from a race detail response.
    If winners_only=True, returns a list with just the winner.
    If False, returns top 4 placegetters.
    """
    results_order = race_data.get("results", [[]])
    if not results_order or not results_order[0]:
        return []

    runners = race_data.get("runners", [])
    runner_map = {r.get("runnerNumber"): r for r in runners}

    if winners_only == "winners":
        placed_numbers = results_order[0][:1]
    elif winners_only == "top4":
        placed_numbers = results_order[0][:4]
    else:
        placed_numbers = results_order[0]

    results = []
    for position, num in enumerate(placed_numbers, 1):
        runner = runner_map.get(num)
        if not runner:
            continue
        data = extract_runner_data(runner, race_data)
        data["position"] = position
        results.append(data)

    return results


def fetch_all_results(from_date, to_date, jurisdiction, race_type_filter=None,
                      winners_only="winners", venue_filter=None, progress_callback=None):
    """
    Fetch all racing results between two dates for a jurisdiction.
    race_type_filter: "R", "H", "G", or None for all.
    winners_only: "winners", "top4", or "all".
    progress_callback(message, current, total) is called to report progress.
    """
    all_rows = []
    current_date = from_date
    dates = []
    while current_date <= to_date:
        dates.append(current_date)
        current_date += timedelta(days=1)

    if not dates:
        return all_rows

    total_dates = len(dates)
    total_meetings = 0

    # First pass: collect all meetings across all dates
    all_meetings_by_date = {}
    for i, d in enumerate(dates):
        date_str = d.strftime("%Y-%m-%d")
        if progress_callback:
            progress_callback(f"Scanning {date_str}…", i, total_dates)
        try:
            meetings = fetch_meetings(date_str, jurisdiction, race_type_filter)
            if venue_filter:
                meetings = [m for m in meetings if m.get("meetingName", "") in venue_filter]
            all_meetings_by_date[date_str] = meetings
            total_meetings += len(meetings)
        except requests.exceptions.HTTPError as e:
            if e.response is not None and e.response.status_code == 404:
                all_meetings_by_date[date_str] = []
            else:
                raise
        except Exception:
            all_meetings_by_date[date_str] = []

    if total_meetings == 0:
        return all_rows

    # Second pass: fetch race details for each meeting
    meeting_count = 0
    for date_str, meetings in all_meetings_by_date.items():
        for meeting in meetings:
            meeting_name = meeting.get("meetingName", "Unknown")
            meeting_count += 1

            if progress_callback:
                progress_callback(
                    f"Fetching {meeting_name}… {meeting_count}/{total_meetings} meetings done",
                    meeting_count,
                    total_meetings,
                )

            races = meeting.get("races", [])
            for race in races:
                race_link = race.get("_links", {}).get("self")
                if not race_link:
                    continue

                try:
                    race_data = fetch_race_detail(race_link)
                except Exception:
                    continue

                race_number = race_data.get("raceNumber", "")
                race_name = race_data.get("raceName", "")
                distance = race_data.get("raceDistance", "")

                results = extract_results(race_data, winners_only=winners_only)
                if not results:
                    continue

                for result in results:
                    pos = result["position"]
                    label = "1st" if pos == 1 else f"{pos}{'nd' if pos == 2 else 'rd' if pos == 3 else 'th'}"
                    all_rows.append({
                        "Date": date_str,
                        "Venue": meeting_name,
                        "Race": race_number,
                        "Race Name": race_name,
                        "Distance": distance,
                        "Position": label,
                        "Runner": result["runner_name"],
                        "No.": result["runner_number"],
                        "Barrier": result["barrier"],
                        "Jockey": result["jockey"],
                        "Trainer": result["trainer"],
                        "Fixed Win": result["fixed_win"],
                        "Tote Win": result["pari_win"],
                    })

    return all_rows


def export_to_xlsx(rows, filepath, selected_columns=None):
    """Export result rows to a formatted .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    all_cols = [
        "Date", "Venue", "Race", "Race Name", "Distance",
        "Position", "Runner", "No.", "Barrier", "Jockey", "Trainer",
        "Fixed Win", "Tote Win",
    ]
    if selected_columns:
        columns = [c for c in all_cols if c in selected_columns]
    else:
        columns = all_cols

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_name in ("Race", "No.", "Barrier", "Distance"):
                cell.alignment = Alignment(horizontal="center")
            if col_name in ("Fixed Win", "Tote Win") and value is not None:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(filepath)


def open_file(filepath):
    """Open a file with the default system application."""
    if platform.system() == "Windows":
        os.startfile(filepath)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", filepath])
    else:
        subprocess.Popen(["xdg-open", filepath])


# ── GUI ─────────────────────────────────────────────────────────────────────

class DateEntry(ttk.Frame):
    """Simple date picker with day/month/year spinboxes."""

    def __init__(self, parent, default_date=None):
        super().__init__(parent)
        if default_date is None:
            default_date = datetime.today()

        self.day_var = tk.StringVar(value=str(default_date.day).zfill(2))
        self.month_var = tk.StringVar(value=str(default_date.month).zfill(2))
        self.year_var = tk.StringVar(value=str(default_date.year))

        vcmd_day = (self.register(lambda v: self._validate_num(v, 1, 31)), "%P")
        vcmd_month = (self.register(lambda v: self._validate_num(v, 1, 12)), "%P")
        vcmd_year = (self.register(lambda v: self._validate_num(v, 2000, 2099)), "%P")

        self.day_spin = ttk.Spinbox(
            self, from_=1, to=31, width=3, textvariable=self.day_var,
            format="%02.0f", validate="focusout", validatecommand=vcmd_day,
        )
        ttk.Label(self, text="/").pack(side=tk.LEFT)
        self.day_spin.pack(side=tk.LEFT)
        ttk.Label(self, text="/").pack(side=tk.LEFT)

        self.month_spin = ttk.Spinbox(
            self, from_=1, to=12, width=3, textvariable=self.month_var,
            format="%02.0f", validate="focusout", validatecommand=vcmd_month,
        )
        self.month_spin.pack(side=tk.LEFT)
        ttk.Label(self, text="/").pack(side=tk.LEFT)

        self.year_spin = ttk.Spinbox(
            self, from_=2000, to=2099, width=5, textvariable=self.year_var,
            validate="focusout", validatecommand=vcmd_year,
        )
        self.year_spin.pack(side=tk.LEFT)

    @staticmethod
    def _validate_num(value, lo, hi):
        if value == "":
            return True
        try:
            n = int(value)
            return lo <= n <= hi
        except ValueError:
            return False

    def get_date(self):
        """Return the selected date as a datetime.date, or None if invalid."""
        try:
            return datetime(
                int(self.year_var.get()),
                int(self.month_var.get()),
                int(self.day_var.get()),
            ).date()
        except (ValueError, TypeError):
            return None


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sports Bet Extractor")
        self.resizable(False, False)
        self.configure(padx=20, pady=15)

        # Determine default save location (Desktop)
        if platform.system() == "Windows":
            self.save_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        else:
            self.save_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.isdir(self.save_dir):
            self.save_dir = os.path.expanduser("~")

        self._build_ui()

        # Centre window on screen
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"+{x}+{y}")

    def _build_ui(self):
        today = datetime.today()
        row = 0

        # ── Title ───────────────────────────────────────────────────────
        title_label = tk.Label(
            self, text="Sports Bet Extractor", font=("Segoe UI", 16, "bold"),
        )
        title_label.grid(row=row, column=0, columnspan=4, pady=(0, 15))
        row += 1

        # ── Date pickers ────────────────────────────────────────────────
        ttk.Label(self, text="From:").grid(row=row, column=0, sticky="e", padx=(0, 5))
        self.from_date = DateEntry(self, default_date=today)
        self.from_date.grid(row=row, column=1, sticky="w")

        ttk.Label(self, text="To:").grid(row=row, column=2, sticky="e", padx=(15, 5))
        self.to_date = DateEntry(self, default_date=today)
        self.to_date.grid(row=row, column=3, sticky="w")
        row += 1

        # ── Jurisdiction ────────────────────────────────────────────────
        ttk.Label(self, text="State:").grid(row=row, column=0, sticky="e", padx=(0, 5), pady=(10, 0))
        self.jurisdiction_var = tk.StringVar(value="QLD")
        jurisdiction_combo = ttk.Combobox(
            self, textvariable=self.jurisdiction_var, values=JURISDICTIONS,
            state="readonly", width=8,
        )
        jurisdiction_combo.grid(row=row, column=1, sticky="w", pady=(10, 0))

        # ── Race type ──────────────────────────────────────────────────
        ttk.Label(self, text="Race type:").grid(row=row, column=2, sticky="e", padx=(15, 5), pady=(10, 0))
        self.race_type_var = tk.StringVar(value="Thoroughbreds")
        race_type_combo = ttk.Combobox(
            self, textvariable=self.race_type_var,
            values=list(RACE_TYPES.keys()), state="readonly", width=14,
        )
        race_type_combo.grid(row=row, column=3, sticky="w", pady=(10, 0))
        row += 1

        # ── Venues filter ──────────────────────────────────────────────
        ttk.Label(self, text="Venues:").grid(row=row, column=0, sticky="e", padx=(0, 5), pady=(10, 0))
        self.venue_filter_var = tk.StringVar(value="All venues")
        venue_combo = ttk.Combobox(
            self, textvariable=self.venue_filter_var,
            values=list(VENUE_FILTERS.keys()), state="readonly", width=14,
        )
        venue_combo.grid(row=row, column=1, sticky="w", pady=(10, 0))
        row += 1

        # ── Results scope ──────────────────────────────────────────────
        ttk.Label(self, text="Show:").grid(row=row, column=0, sticky="e", padx=(0, 5), pady=(10, 0))
        self.scope_var = tk.StringVar(value="Winners only")
        scope_combo = ttk.Combobox(
            self, textvariable=self.scope_var,
            values=["Winners only", "Top 4 placegetters", "All runners"],
            state="readonly", width=18,
        )
        scope_combo.grid(row=row, column=1, columnspan=2, sticky="w", pady=(10, 0))
        row += 1

        # ── Column checkboxes ──────────────────────────────────────────
        col_frame = ttk.LabelFrame(self, text="Columns to include", padding=(10, 5))
        col_frame.grid(row=row, column=0, columnspan=4, sticky="ew", pady=(10, 0))

        self.column_vars = {}
        col_names = [
            "Date", "Venue", "Race", "Race Name", "Distance",
            "Position", "Runner", "No.", "Barrier", "Jockey", "Trainer",
            "Fixed Win", "Tote Win",
        ]
        cols_per_row = 5
        for i, col_name in enumerate(col_names):
            var = tk.BooleanVar(value=True)
            self.column_vars[col_name] = var
            cb = ttk.Checkbutton(col_frame, text=col_name, variable=var)
            cb.grid(row=i // cols_per_row, column=i % cols_per_row, sticky="w", padx=(0, 10))

        # Select All / Deselect All buttons
        btn_frame = ttk.Frame(col_frame)
        btn_frame.grid(row=(len(col_names) // cols_per_row) + 1, column=0,
                       columnspan=cols_per_row, sticky="w", pady=(5, 0))
        ttk.Button(btn_frame, text="Select All", command=self._select_all_cols, width=10).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_frame, text="Deselect All", command=self._deselect_all_cols, width=10).pack(side=tk.LEFT)
        row += 1

        # ── Save location ──────────────────────────────────────────────
        ttk.Label(self, text="Save to:").grid(row=row, column=0, sticky="e", padx=(0, 5), pady=(10, 0))

        save_frame = ttk.Frame(self)
        save_frame.grid(row=row, column=1, columnspan=3, sticky="ew", pady=(10, 0))

        self.save_dir_var = tk.StringVar(value=self.save_dir)
        self.save_label = ttk.Label(
            save_frame, textvariable=self.save_dir_var, width=40,
            relief="sunken", padding=(5, 2),
        )
        self.save_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        browse_btn = ttk.Button(save_frame, text="Browse…", command=self._browse, width=8)
        browse_btn.pack(side=tk.LEFT, padx=(5, 0))
        row += 1

        # ── Open file checkbox ──────────────────────────────────────────
        self.open_after_var = tk.BooleanVar(value=True)
        open_cb = ttk.Checkbutton(
            self, text="Open file after download", variable=self.open_after_var,
        )
        open_cb.grid(row=row, column=1, columnspan=2, sticky="w", pady=(10, 0))
        row += 1

        # ── Download button ─────────────────────────────────────────────
        self.download_btn = ttk.Button(
            self, text="Download Results", command=self._on_download,
        )
        self.download_btn.grid(row=row, column=0, columnspan=4, pady=(15, 10), ipadx=20, ipady=5)
        row += 1

        # ── Status bar ──────────────────────────────────────────────────
        self.status_var = tk.StringVar(value="Ready")
        self.status_label = tk.Label(
            self, textvariable=self.status_var, anchor="w",
            font=("Segoe UI", 9), fg="#555555", relief="sunken",
            padx=8, pady=4,
        )
        self.status_label.grid(row=row, column=0, columnspan=4, sticky="ew", pady=(5, 0))

    def _select_all_cols(self):
        for var in self.column_vars.values():
            var.set(True)

    def _deselect_all_cols(self):
        for var in self.column_vars.values():
            var.set(False)

    def _browse(self):
        chosen = filedialog.askdirectory(initialdir=self.save_dir_var.get())
        if chosen:
            self.save_dir_var.set(chosen)

    def _set_status(self, message, colour="#555555"):
        self.status_var.set(message)
        self.status_label.configure(fg=colour)

    def _get_selected_columns(self):
        return {name for name, var in self.column_vars.items() if var.get()}

    def _on_download(self):
        # Validate dates
        from_d = self.from_date.get_date()
        to_d = self.to_date.get_date()

        if from_d is None or to_d is None:
            messagebox.showerror("Invalid Date", "Please enter valid dates in DD/MM/YYYY format.")
            return

        if from_d > to_d:
            messagebox.showerror("Invalid Date", "'From' date must be on or before 'To' date.")
            return

        selected_cols = self._get_selected_columns()
        if not selected_cols:
            messagebox.showerror("No Columns", "Please select at least one column to include.")
            return

        jurisdiction = self.jurisdiction_var.get()
        save_dir = self.save_dir_var.get()

        if not os.path.isdir(save_dir):
            messagebox.showerror("Invalid Folder", "The save location does not exist.")
            return

        race_type_filter = RACE_TYPES.get(self.race_type_var.get())
        venue_filter = VENUE_FILTERS.get(self.venue_filter_var.get())
        scope = self.scope_var.get()
        if scope == "Winners only":
            scope_mode = "winners"
        elif scope == "Top 4 placegetters":
            scope_mode = "top4"
        else:
            scope_mode = "all"

        filename = (
            f"sportsbetresults_{from_d.strftime('%Y-%m-%d')}"
            f"_to_{to_d.strftime('%Y-%m-%d')}_{jurisdiction}.xlsx"
        )
        filepath = os.path.join(save_dir, filename)

        # Disable button
        self.download_btn.configure(state="disabled", text="Downloading…")
        self._set_status("Starting…")

        # Run in background thread
        thread = threading.Thread(
            target=self._download_thread,
            args=(from_d, to_d, jurisdiction, filepath, race_type_filter,
                  scope_mode, selected_cols, venue_filter),
            daemon=True,
        )
        thread.start()

    def _download_thread(self, from_d, to_d, jurisdiction, filepath,
                         race_type_filter, scope_mode, selected_cols, venue_filter):
        def progress(msg, current, total):
            self.after(0, self._set_status, msg)

        try:
            rows = fetch_all_results(
                from_d, to_d, jurisdiction,
                race_type_filter=race_type_filter,
                winners_only=scope_mode,
                venue_filter=venue_filter,
                progress_callback=progress,
            )

            if not rows:
                self.after(0, self._finish_download, None, "No results found for the selected dates and state.", True)
                return

            export_to_xlsx(rows, filepath, selected_columns=selected_cols)
            self.after(0, self._finish_download, filepath, None, False)

        except requests.exceptions.ConnectionError:
            self.after(
                0, self._finish_download, None,
                "Connection error — check your internet connection and try again.", True,
            )
        except requests.exceptions.Timeout:
            self.after(
                0, self._finish_download, None,
                "Request timed out — the TAB API may be slow. Try again.", True,
            )
        except Exception as e:
            self.after(
                0, self._finish_download, None,
                f"Error: {e}", True,
            )

    def _finish_download(self, filepath, error_msg, is_error):
        self.download_btn.configure(state="normal", text="Download Results")

        if is_error:
            self._set_status(error_msg, colour="#CC0000")
            return

        short_path = filepath
        try:
            short_path = os.path.basename(filepath)
        except Exception:
            pass

        self._set_status(f"Done — saved to {short_path}", colour="#007A33")

        if self.open_after_var.get() and filepath:
            try:
                open_file(filepath)
            except Exception:
                pass


# ── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
